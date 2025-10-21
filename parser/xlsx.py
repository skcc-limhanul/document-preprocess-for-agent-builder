from openpyxl import load_workbook


# def parse_content(file_path):
#     '''
#     Extract contents from xlsx including merged cells, collapsing them to show only the first value.
#     Parameter:
#     * file_path : The path to xlsx
#     Return:
#     * parsed_text : Parsed contents like plain texts and markdown style tables
#     '''
#     parsed_text = ''
#     wb = load_workbook(file_path, data_only=True)
#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         parsed_text += f"\n### Sheet: {sheet_name}\n\n"

#         merged_cell_map = {}
#         for merged_range in ws.merged_cells.ranges:
#             min_row, min_col = merged_range.min_row, merged_range.min_col
#             value = ws.cell(row=min_row, column=min_col).value
#             for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row,
#                                     min_col=merged_range.min_col, max_col=merged_range.max_col):
#                 for cell in row:
#                     merged_cell_map[(cell.row, cell.column)] = value

#         for row in ws.iter_rows():
#             row_text = []
#             for cell in row:
#                 val = cell.value
#                 if val is None and (cell.row, cell.column) in merged_cell_map:
#                     val = merged_cell_map[(cell.row, cell.column)]
#                 row_text.append(str(val).strip() if val else '')
            
#             if all(cell == '' for cell in row_text):
#                 parsed_text += '\n'
#                 continue

#             non_empty_count = sum(1 for cell in row_text if cell != '')
#             if non_empty_count >= 3:
#                 parsed_text += '| ' + ' | '.join(row_text) + ' |\n'
#             else:
#                 parsed_text += ' '.join(row_text) + '\n'
#     return parsed_text




def parse_content(file_path):
    '''
    Extract contents from xlsx including merged cells, collapsing them to show only the first value.
    Detect tables by continuous rows with multiple non-empty cells.
    Output markdown-style tables and plain text lines.
    '''
    parsed_text = ''
    wb = load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parsed_text += f"\n### Sheet: {sheet_name}\n\n"

        # Build merged cell map
        merged_cell_map = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            value = ws.cell(row=min_row, column=min_col).value
            for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row,
                                    min_col=merged_range.min_col, max_col=merged_range.max_col):
                for cell in row:
                    merged_cell_map[(cell.row, cell.column)] = value

        # Function to trim trailing empty cells in a row
        def trim_row(row):
            while row and row[-1] == '':
                row.pop()
            return row

        table_buffer = []
        inside_table = False

        for row in ws.iter_rows():
            row_values = []
            for cell in row:
                val = cell.value
                if val is None and (cell.row, cell.column) in merged_cell_map:
                    val = merged_cell_map[(cell.row, cell.column)]
                # Convert to string, strip spaces, or empty string if None
                val_str = str(val).strip() if val is not None else ''
                row_values.append(val_str)

            # Trim trailing empty columns
            row_values = trim_row(row_values)

            # Count non-empty cells (excluding empty strings)
            non_empty_count = sum(1 for c in row_values if c != '')

            # Decide if this row is part of a table
            # Threshold: at least 2 non-empty cells (adjust if needed)
            if non_empty_count >= 2:
                # Start or continue table
                table_buffer.append(row_values)
                inside_table = True
            else:
                # Row has fewer non-empty cells: close table if open
                if inside_table:
                    # Output the buffered table as markdown
                    if len(table_buffer) > 0:
                        # Find max number of columns in the table
                        max_cols = max(len(r) for r in table_buffer)

                        # Normalize all rows to max_cols
                        norm_rows = [r + ['']*(max_cols - len(r)) for r in table_buffer]

                        # Output header
                        header = norm_rows[0]
                        parsed_text += '| ' + ' | '.join(header) + ' |\n'
                        # Output separator
                        parsed_text += '| ' + ' | '.join(['---'] * max_cols) + ' |\n'

                        # Output data rows
                        for data_row in norm_rows[1:]:
                            parsed_text += '| ' + ' | '.join(data_row) + ' |\n'

                    parsed_text += '\n'
                    table_buffer = []
                    inside_table = False

                # Output this non-table row as plain text if it has any content
                if any(c != '' for c in row_values):
                    parsed_text += ' '.join(row_values) + '\n'
                else:
                    parsed_text += '\n'

        # If file ends inside a table, output it too
        if inside_table and len(table_buffer) > 0:
            max_cols = max(len(r) for r in table_buffer)
            norm_rows = [r + ['']*(max_cols - len(r)) for r in table_buffer]

            header = norm_rows[0]
            parsed_text += '| ' + ' | '.join(header) + ' |\n'
            parsed_text += '| ' + ' | '.join(['---'] * max_cols) + ' |\n'

            for data_row in norm_rows[1:]:
                parsed_text += '| ' + ' | '.join(data_row) + ' |\n'
            parsed_text += '\n'

    return parsed_text
