from openpyxl import load_workbook


def _extract_text_from_xlsx(file_path):
    '''
    Extract contents include texts from contexts and tables from table shaped contexts.\n
    It can extract contents from multiple sheets of xlsx file. And no need to un-merge cells.
    Parameter
    * file_path : The path of xlsx to extract contents
    Return
    * parsed_text : Parsed contents like plain texts and markdown style tables
    '''
    parsed_text = ''
    wb = load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parsed_text = parsed_text + '\n' + f"\n📄 Sheet: {sheet_name}"
        for row in ws.iter_rows(values_only=True):
            row_text =\
                [str(cell).strip() if cell is not None else '' for cell in row]
            parsed_text = parsed_text + '\n' + ' | '.join(row_text) + ' |'
    return parsed_text
